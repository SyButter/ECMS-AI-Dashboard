"""
ECMS Excel Import Script
========================
Reads the ECMS Excel workbook and loads data into MySQL.

Sheets handled:
  - Summary Level  → programs table
  - Data Level     → firms + program_firms tables

Usage:
  python scripts/import_excel.py --file path/to/ECMS_data.xlsx
  python scripts/import_excel.py --file path/to/ECMS_data.xlsx --dry-run

Options:
  --file      Path to the Excel file (required)
  --dry-run   Parse and validate without writing to the database
  --verbose   Print every row as it is processed

Requirements:
  pip install openpyxl python-dotenv mysql-connector-python

Environment (.env file):
  DB_HOST, DB_PORT, DB_NAME, DB_USER, DB_PASSWORD
"""

import argparse
import logging
import os
import sys
from dataclasses import dataclass, field
from datetime import datetime, date
from typing import Optional

import mysql.connector
from mysql.connector import MySQLConnection
import openpyxl
from dotenv import load_dotenv

# ---------------------------------------------------------------------------
# Logging — writes to console AND to logs/import_YYYYMMDD_HHMMSS.log
# Force UTF-8 on all handlers so Windows CP1252 consoles don't choke
# on special characters in warning messages.
# ---------------------------------------------------------------------------
os.makedirs("logs", exist_ok=True)
log_filename = f"logs/import_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

# Console handler — UTF-8 forced, falls back gracefully on older Windows
_console_handler = logging.StreamHandler(
    stream=open(sys.stdout.fileno(), mode="w", encoding="utf-8", buffering=1, closefd=False)
    if hasattr(sys.stdout, "fileno")
    else sys.stdout
)
_console_handler.setFormatter(logging.Formatter("%(asctime)s  %(levelname)-8s  %(message)s"))

# File handler — always UTF-8
_file_handler = logging.FileHandler(log_filename, encoding="utf-8")
_file_handler.setFormatter(logging.Formatter("%(asctime)s  %(levelname)-8s  %(message)s"))

logging.basicConfig(level=logging.INFO, handlers=[_console_handler, _file_handler])
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Run summary — tracks what happened across the whole import
# ---------------------------------------------------------------------------
@dataclass
class RunSummary:
    sheet: str = ""
    rows_read: int = 0
    inserted: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list = field(default_factory=list)

    def report(self):
        log.info("=" * 60)
        log.info(f"  Sheet     : {self.sheet}")
        log.info(f"  Rows read : {self.rows_read}")
        log.info(f"  Inserted  : {self.inserted}")
        log.info(f"  Updated   : {self.updated}")
        log.info(f"  Skipped   : {self.skipped}")
        log.info(f"  Errors    : {len(self.errors)}")
        if self.errors:
            for e in self.errors:
                log.warning(f"    [!] {e}")
        log.info("=" * 60)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def cell_val(row, idx: int, default=None):
    """Safely get a cell value from a row tuple by 0-based index."""
    try:
        v = row[idx]
        if v is None or str(v).strip() == "":
            return default
        return str(v).strip() if isinstance(v, str) else v
    except IndexError:
        return default


def to_date(value) -> Optional[date]:
    """Convert various date formats from Excel to Python date."""
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value if isinstance(value, date) else value.date()
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def to_float(value) -> Optional[float]:
    """Convert string or number to float, strip $ and commas."""
    if value is None:
        return None
    try:
        return float(str(value).replace("$", "").replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def to_bool(value) -> int:
    """Convert Yes/No/1/0/True/False to MySQL tinyint."""
    if value is None:
        return 0
    s = str(value).strip().lower()
    return 1 if s in ("yes", "y", "1", "true", "x") else 0


# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------

def get_connection() -> MySQLConnection:
    load_dotenv()
    return mysql.connector.connect(
        host=os.getenv("DB_HOST", "localhost"),
        port=int(os.getenv("DB_PORT", 3306)),
        database=os.getenv("DB_NAME", "ecms"),
        user=os.getenv("DB_USER", "root"),
        password=os.getenv("DB_PASSWORD", ""),
        autocommit=False,
    )


def get_or_create_division(cursor, code: str, name: str = None) -> Optional[int]:
    """Return division_id for code, creating it if it doesn't exist."""
    if not code:
        return None
    cursor.execute("SELECT division_id FROM divisions WHERE division_code = %s", (code,))
    row = cursor.fetchone()
    if row:
        return row[0]
    # Create it — name defaults to code if not provided
    cursor.execute(
        "INSERT INTO divisions (division_code, division_name) VALUES (%s, %s)",
        (code, name or code),
    )
    log.info(f"  Created new division: {code}")
    return cursor.lastrowid


def get_or_create_unit(cursor, division_id: int, unit_name: str) -> Optional[int]:
    """Return unit_id, creating the unit if it doesn't exist."""
    if not unit_name or not division_id:
        return None
    cursor.execute(
        "SELECT unit_id FROM units WHERE division_id = %s AND unit_name = %s",
        (division_id, unit_name),
    )
    row = cursor.fetchone()
    if row:
        return row[0]
    cursor.execute(
        "INSERT INTO units (division_id, unit_name) VALUES (%s, %s)",
        (division_id, unit_name),
    )
    log.info(f"  Created new unit: {unit_name}")
    return cursor.lastrowid


def get_or_create_solicitation_type(cursor, label: str) -> Optional[int]:
    """Return type_id for label, creating it if needed."""
    if not label:
        return None
    cursor.execute(
        "SELECT type_id FROM solicitation_types WHERE type_label = %s", (label,)
    )
    row = cursor.fetchone()
    if row:
        return row[0]
    code = label.lower().replace(" ", "_").replace("-", "_")
    cursor.execute(
        "INSERT INTO solicitation_types (type_code, type_label) VALUES (%s, %s)",
        (code, label),
    )
    log.info(f"  Created new solicitation type: {label}")
    return cursor.lastrowid


def get_or_create_category(cursor, name: str) -> Optional[int]:
    """Return category_id, creating if needed."""
    if not name:
        return None
    cursor.execute(
        "SELECT category_id FROM categories WHERE category_name = %s", (name,)
    )
    row = cursor.fetchone()
    if row:
        return row[0]
    cursor.execute(
        "INSERT INTO categories (category_name) VALUES (%s)", (name,)
    )
    log.info(f"  Created new category: {name}")
    return cursor.lastrowid


def get_or_create_firm(cursor, firm_name: str, row_data: dict) -> Optional[int]:
    """
    Return firm_id for a firm name.
    If not found, insert a new firm record with whatever data is available.
    Matches on firm_name (case-insensitive) or global_vendor_id.
    """
    if not firm_name:
        return None

    # Try Global Vendor ID first (most reliable match)
    gvid = row_data.get("global_vendor_id")
    if gvid:
        cursor.execute(
            "SELECT firm_id FROM firms WHERE global_vendor_id = %s", (gvid,)
        )
        row = cursor.fetchone()
        if row:
            return row[0]

    # Fall back to name match
    cursor.execute(
        "SELECT firm_id FROM firms WHERE LOWER(firm_name) = LOWER(%s)", (firm_name,)
    )
    row = cursor.fetchone()
    if row:
        return row[0]

    # Create new firm
    cursor.execute(
        """
        INSERT INTO firms (
            firm_name, alternate_name, global_vendor_id, sap_vendor_number,
            is_sbe, is_wbe, is_dbe, is_mbe, is_mwbe, is_sdvob, is_ai, is_lbe, is_certified,
            primary_contact_name, primary_contact_title, primary_contact_email,
            secondary_contact_name, secondary_contact_email,
            third_contact_name, third_contact_email,
            address1, city, state, zip
        ) VALUES (
            %s, %s, %s, %s,
            %s, %s, %s, %s, %s, %s, %s, %s, %s,
            %s, %s, %s,
            %s, %s,
            %s, %s,
            %s, %s, %s, %s
        )
        """,
        (
            firm_name,
            row_data.get("alternate_name"),
            row_data.get("global_vendor_id"),
            row_data.get("sap_vendor_number"),
            row_data.get("is_sbe", 0),
            row_data.get("is_wbe", 0),
            row_data.get("is_dbe", 0),
            row_data.get("is_mbe", 0),
            row_data.get("is_mwbe", 0),
            row_data.get("is_sdvob", 0),
            row_data.get("is_ai", 0),
            row_data.get("is_lbe", 0),
            row_data.get("is_certified", 0),
            row_data.get("primary_contact_name"),
            row_data.get("primary_contact_title"),
            row_data.get("primary_contact_email"),
            row_data.get("secondary_contact_name"),
            row_data.get("secondary_contact_email"),
            row_data.get("third_contact_name"),
            row_data.get("third_contact_email"),
            row_data.get("address1"),
            row_data.get("city"),
            row_data.get("state"),
            row_data.get("zip"),
        ),
    )
    log.info(f"  Created new firm: {firm_name}")
    return cursor.lastrowid


# ---------------------------------------------------------------------------
# SHEET 1: Summary Level  →  programs table
# ---------------------------------------------------------------------------
# Expected columns (0-based index → field name):
#   0  Division
#   1  Division #
#   2  Unit / Unit Group
#   3  RFP #
#   4  Program Type
#   5  Program Title
#   6  Status
#   7  Solicitation Date
#   8  Category
#   9  PA Contact
#   10 Type (award type)
#   11 ICE ($)
#   12 # Firms on List
#   13 # Submissions
#   14 Awarded Firm
#   15 Agreement No.
#   16 Award Date
#   17 Award Amount ($)
#   18 PO #
# ---------------------------------------------------------------------------

SUMMARY_COL = {
    "division_code":      0,
    "division_number":    1,
    "unit_name":          2,
    "rfp_number":         3,
    "program_type":       4,
    "program_title":      5,
    "status":             6,
    "solicitation_date":  7,
    "category":           8,
    "pa_contact":         9,
    "award_type":         10,
    "ice_estimate":       11,
    "num_firms_on_list":  12,
    "num_submissions":    13,
    "awarded_firm":       14,
    "agreement_number":   15,
    "award_date":         16,
    "award_amount":       17,
    "po_number":          18,
}


def import_summary_level(ws, cursor, summary: RunSummary, dry_run: bool, verbose: bool):
    """Process the Summary Level sheet → programs table."""
    log.info("Processing Summary Level sheet...")

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip completely empty rows
        if all(v is None for v in row):
            continue

        summary.rows_read += 1
        rfp_number = cell_val(row, SUMMARY_COL["rfp_number"])

        # RFP # is the primary key for a program — skip if missing
        if not rfp_number:
            summary.skipped += 1
            summary.errors.append(f"Row {row_idx}: Missing RFP # — skipped")
            continue

        if verbose:
            log.info(f"  Row {row_idx}: RFP {rfp_number}")

        try:
            # Resolve foreign keys
            div_code = cell_val(row, SUMMARY_COL["division_code"])
            div_num  = cell_val(row, SUMMARY_COL["division_number"])
            division_id = get_or_create_division(cursor, div_code, div_code) if div_code else None

            unit_name  = cell_val(row, SUMMARY_COL["unit_name"])
            unit_id    = get_or_create_unit(cursor, division_id, unit_name) if (division_id and unit_name) else None

            prog_type  = cell_val(row, SUMMARY_COL["program_type"])
            type_id    = get_or_create_solicitation_type(cursor, prog_type) if prog_type else None

            category   = cell_val(row, SUMMARY_COL["category"])
            category_id = get_or_create_category(cursor, category) if category else None

            # Build program record
            program_title    = cell_val(row, SUMMARY_COL["program_title"]) or rfp_number
            status_raw       = cell_val(row, SUMMARY_COL["status"], "Active")
            status           = status_raw if status_raw in ("Active","Closed","Pending","Cancelled") else "Active"
            solicitation_date = to_date(cell_val(row, SUMMARY_COL["solicitation_date"]))
            award_date        = to_date(cell_val(row, SUMMARY_COL["award_date"]))
            ice_estimate      = to_float(cell_val(row, SUMMARY_COL["ice_estimate"]))
            num_firms         = cell_val(row, SUMMARY_COL["num_firms_on_list"])
            num_subs          = cell_val(row, SUMMARY_COL["num_submissions"])
            agreement_number  = cell_val(row, SUMMARY_COL["agreement_number"])
            pa_contact        = cell_val(row, SUMMARY_COL["pa_contact"])
            program_type_str  = cell_val(row, SUMMARY_COL["program_type"])

            if not dry_run:
                # Upsert: update if RFP # already exists, insert if not
                cursor.execute(
                    "SELECT program_id FROM programs WHERE rfp_number = %s", (rfp_number,)
                )
                existing = cursor.fetchone()

                if existing:
                    cursor.execute(
                        """
                        UPDATE programs SET
                            program_title        = %s,
                            agreement_number     = %s,
                            solicitation_type_id = COALESCE(%s, solicitation_type_id),
                            division_id          = COALESCE(%s, division_id),
                            unit_id              = COALESCE(%s, unit_id),
                            ice_estimate         = COALESCE(%s, ice_estimate),
                            num_firms_on_list    = COALESCE(%s, num_firms_on_list),
                            num_submissions      = COALESCE(%s, num_submissions),
                            solicitation_date    = COALESCE(%s, solicitation_date),
                            award_date           = COALESCE(%s, award_date),
                            pa_contact           = COALESCE(%s, pa_contact),
                            status               = %s,
                            program_type         = COALESCE(%s, program_type),
                            updated_at           = NOW()
                        WHERE rfp_number = %s
                        """,
                        (
                            program_title, agreement_number,
                            type_id, division_id, unit_id,
                            ice_estimate, num_firms, num_subs,
                            solicitation_date, award_date,
                            pa_contact, status, program_type_str,
                            rfp_number,
                        ),
                    )
                    program_id = existing[0]
                    summary.updated += 1
                else:
                    cursor.execute(
                        """
                        INSERT INTO programs (
                            rfp_number, program_title, agreement_number,
                            solicitation_type_id, division_id, unit_id,
                            ice_estimate, num_firms_on_list, num_submissions,
                            solicitation_date, award_date,
                            pa_contact, status, program_type
                        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        """,
                        (
                            rfp_number, program_title, agreement_number,
                            type_id, division_id, unit_id,
                            ice_estimate, num_firms, num_subs,
                            solicitation_date, award_date,
                            pa_contact, status, program_type_str,
                        ),
                    )
                    program_id = cursor.lastrowid
                    summary.inserted += 1

                # Link category to program
                if category_id and program_id:
                    cursor.execute(
                        """
                        INSERT IGNORE INTO program_categories (program_id, category_id, sort_order)
                        VALUES (%s, %s, 1)
                        """,
                        (program_id, category_id),
                    )

                # If there's an awarded firm on this row, ensure it exists in firms
                awarded_firm = cell_val(row, SUMMARY_COL["awarded_firm"])
                if awarded_firm:
                    firm_id = get_or_create_firm(cursor, awarded_firm, {})
                    # Link as program_firm if not already linked
                    if firm_id:
                        cursor.execute(
                            """
                            INSERT IGNORE INTO program_firms (program_id, firm_id, date_added)
                            VALUES (%s, %s, %s)
                            """,
                            (program_id, firm_id, award_date or solicitation_date),
                        )

        except Exception as e:
            summary.errors.append(f"Row {row_idx} (RFP {rfp_number}): {e}")
            log.error(f"  ✗ Row {row_idx}: {e}")
            continue


# ---------------------------------------------------------------------------
# SHEET 2: Data Level  →  firms + program_firms tables
# ---------------------------------------------------------------------------
# Expected columns (0-based index → field name):
#   0  AgmtFirmRecord (Yes/No)
#   1  RFP Shorthand
#   2  Date Valid
#   3  RFP No.
#   4  Program Type
#   5  Agreement No.
#   6  Firm Name
#   7  Category 1
#   8  Category 2
#   9  Category 3
#   10 Category 4
#   11 Division
#   12 Unit_Group
#   13 Alternate Name
#   14 SCP List (Yes/No)
#   15 SBE
#   16 WBE
#   17 DBE
#   18 MBE
#   19 MWBE
#   20 SDVOB
#   21 AI (Asian Indian)
#   22 LBE
#   23 Is Certified
#   24 Global Vendor ID
#   25 SAP Vendor #
#   26 Salutation
#   27 Primary Contact Name
#   28 Primary Contact Title
#   29 Primary Contact Email
#   30 Secondary Contact Name
#   31 Secondary Contact Email
#   32 Third Contact Name
#   33 Third Contact Email
#   34 Address 1
#   35 City
#   36 State
#   37 Zip
#   38 Security Level
#   39 Date Sent to Procurement
#   40 PM Name
#   41 Agree PM
#   42 DAR Name
# ---------------------------------------------------------------------------

DATA_COL = {
    "agmt_firm_record":       0,
    "rfp_shorthand":          1,
    "date_valid":             2,
    "rfp_number":             3,
    "program_type":           4,
    "agreement_number":       5,
    "firm_name":              6,
    "category_1":             7,
    "category_2":             8,
    "category_3":             9,
    "category_4":             10,
    "division_code":          11,
    "unit_name":              12,
    "alternate_name":         13,
    "scp_list":               14,
    "is_sbe":                 15,
    "is_wbe":                 16,
    "is_dbe":                 17,
    "is_mbe":                 18,
    "is_mwbe":                19,
    "is_sdvob":               20,
    "is_ai":                  21,
    "is_lbe":                 22,
    "is_certified":           23,
    "global_vendor_id":       24,
    "sap_vendor_number":      25,
    "salutation":             26,
    "primary_contact_name":   27,
    "primary_contact_title":  28,
    "primary_contact_email":  29,
    "secondary_contact_name": 35,
    "secondary_contact_email":37,
    "third_contact_name":     38,
    "third_contact_email":    39,
    "address1":               29,
    "city":                   31,
    "state":                  32,
    "zip":                    33,
    "security_level":         41,
    "date_sent_to_procurement":42,
    "pm_name":                43,
    "agree_pm":               44,
    "dar_name":               45,
}


def import_data_level(ws, cursor, summary: RunSummary, dry_run: bool, verbose: bool):
    """Process the Data Level sheet → firms + program_firms tables."""
    log.info("Processing Data Level sheet...")

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue

        summary.rows_read += 1
        firm_name  = cell_val(row, DATA_COL["firm_name"])
        rfp_number = cell_val(row, DATA_COL["rfp_number"])

        if not firm_name:
            summary.skipped += 1
            summary.errors.append(f"Row {row_idx}: Missing Firm Name — skipped")
            continue

        if verbose:
            log.info(f"  Row {row_idx}: {firm_name} / RFP {rfp_number}")

        try:
            # Build firm data dict from this row
            firm_data = {
                "alternate_name":           cell_val(row, DATA_COL["alternate_name"]),
                "global_vendor_id":         cell_val(row, DATA_COL["global_vendor_id"]),
                "sap_vendor_number":        cell_val(row, DATA_COL["sap_vendor_number"]),
                "is_sbe":                   to_bool(cell_val(row, DATA_COL["is_sbe"])),
                "is_wbe":                   to_bool(cell_val(row, DATA_COL["is_wbe"])),
                "is_dbe":                   to_bool(cell_val(row, DATA_COL["is_dbe"])),
                "is_mbe":                   to_bool(cell_val(row, DATA_COL["is_mbe"])),
                "is_mwbe":                  to_bool(cell_val(row, DATA_COL["is_mwbe"])),
                "is_sdvob":                 to_bool(cell_val(row, DATA_COL["is_sdvob"])),
                "is_ai":                    to_bool(cell_val(row, DATA_COL["is_ai"])),
                "is_lbe":                   to_bool(cell_val(row, DATA_COL["is_lbe"])),
                "is_certified":             to_bool(cell_val(row, DATA_COL["is_certified"])),
                "salutation":               cell_val(row, DATA_COL["salutation"]),
                "primary_contact_name":     cell_val(row, DATA_COL["primary_contact_name"]),
                "primary_contact_title":    cell_val(row, DATA_COL["primary_contact_title"]),
                "primary_contact_email":    cell_val(row, DATA_COL["primary_contact_email"]),
                "secondary_contact_name":   cell_val(row, DATA_COL["secondary_contact_name"]),
                "secondary_contact_email":  cell_val(row, DATA_COL["secondary_contact_email"]),
                "third_contact_name":       cell_val(row, DATA_COL["third_contact_name"]),
                "third_contact_email":      cell_val(row, DATA_COL["third_contact_email"]),
                "address1":                 cell_val(row, DATA_COL["address1"]),
                "city":                     cell_val(row, DATA_COL["city"]),
                "state":                    cell_val(row, DATA_COL["state"]),
                "zip":                      cell_val(row, DATA_COL["zip"]),
            }

            if not dry_run:
                # Upsert firm — update certifications and contacts if firm already exists
                gvid = firm_data.get("global_vendor_id")
                existing_firm = None

                if gvid:
                    cursor.execute(
                        "SELECT firm_id FROM firms WHERE global_vendor_id = %s", (gvid,)
                    )
                    existing_firm = cursor.fetchone()

                if not existing_firm:
                    cursor.execute(
                        "SELECT firm_id FROM firms WHERE LOWER(firm_name) = LOWER(%s)",
                        (firm_name,),
                    )
                    existing_firm = cursor.fetchone()

                if existing_firm:
                    firm_id = existing_firm[0]
                    # Update firm details — certifications and contacts may have changed
                    cursor.execute(
                        """
                        UPDATE firms SET
                            alternate_name          = COALESCE(%s, alternate_name),
                            global_vendor_id        = COALESCE(%s, global_vendor_id),
                            sap_vendor_number       = COALESCE(%s, sap_vendor_number),
                            is_sbe                  = %s,
                            is_wbe                  = %s,
                            is_dbe                  = %s,
                            is_mbe                  = %s,
                            is_mwbe                 = %s,
                            is_sdvob                = %s,
                            is_ai                   = %s,
                            is_lbe                  = %s,
                            is_certified            = %s,
                            primary_contact_name    = COALESCE(%s, primary_contact_name),
                            primary_contact_title   = COALESCE(%s, primary_contact_title),
                            primary_contact_email   = COALESCE(%s, primary_contact_email),
                            secondary_contact_name  = COALESCE(%s, secondary_contact_name),
                            secondary_contact_email = COALESCE(%s, secondary_contact_email),
                            third_contact_name      = COALESCE(%s, third_contact_name),
                            third_contact_email     = COALESCE(%s, third_contact_email),
                            address1                = COALESCE(%s, address1),
                            city                    = COALESCE(%s, city),
                            state                   = COALESCE(%s, state),
                            zip                     = COALESCE(%s, zip),
                            updated_at              = NOW()
                        WHERE firm_id = %s
                        """,
                        (
                            firm_data["alternate_name"], firm_data["global_vendor_id"],
                            firm_data["sap_vendor_number"],
                            firm_data["is_sbe"], firm_data["is_wbe"], firm_data["is_dbe"],
                            firm_data["is_mbe"], firm_data["is_mwbe"], firm_data["is_sdvob"],
                            firm_data["is_ai"], firm_data["is_lbe"], firm_data["is_certified"],
                            firm_data["primary_contact_name"], firm_data["primary_contact_title"],
                            firm_data["primary_contact_email"],
                            firm_data["secondary_contact_name"], firm_data["secondary_contact_email"],
                            firm_data["third_contact_name"], firm_data["third_contact_email"],
                            firm_data["address1"], firm_data["city"],
                            firm_data["state"], firm_data["zip"],
                            firm_id,
                        ),
                    )
                    summary.updated += 1
                else:
                    firm_id = get_or_create_firm(cursor, firm_name, firm_data)
                    summary.inserted += 1

                # Link firm to program (program_firms)
                if rfp_number and firm_id:
                    cursor.execute(
                        "SELECT program_id FROM programs WHERE rfp_number = %s",
                        (rfp_number,),
                    )
                    prog_row = cursor.fetchone()
                    if prog_row:
                        program_id = prog_row[0]
                        agmt_flag = to_bool(cell_val(row, DATA_COL["agmt_firm_record"]))
                        scp_flag  = to_bool(cell_val(row, DATA_COL["scp_list"]))
                        date_added = to_date(cell_val(row, DATA_COL["date_valid"]))

                        cursor.execute(
                            """
                            INSERT INTO program_firms (program_id, firm_id, agmt_firm_record,
                                is_on_scp_list, date_added)
                            VALUES (%s, %s, %s, %s, %s)
                            ON DUPLICATE KEY UPDATE
                                agmt_firm_record = VALUES(agmt_firm_record),
                                is_on_scp_list   = VALUES(is_on_scp_list)
                            """,
                            (program_id, firm_id, agmt_flag, scp_flag, date_added),
                        )

                        # Update program-level fields from Data Level sheet
                        rfp_shorthand = cell_val(row, DATA_COL["rfp_shorthand"])
                        pm_name       = cell_val(row, DATA_COL["pm_name"])
                        agree_pm      = cell_val(row, DATA_COL["agree_pm"])
                        dar_name      = cell_val(row, DATA_COL["dar_name"])
                        sec_level     = cell_val(row, DATA_COL["security_level"])
                        date_proc     = to_date(cell_val(row, DATA_COL["date_sent_to_procurement"]))

                        cursor.execute(
                            """
                            UPDATE programs SET
                                rfp_shorthand            = COALESCE(%s, rfp_shorthand),
                                pm_name                  = COALESCE(%s, pm_name),
                                agree_pm                 = COALESCE(%s, agree_pm),
                                dar_name                 = COALESCE(%s, dar_name),
                                security_level           = COALESCE(%s, security_level),
                                date_sent_to_procurement = COALESCE(%s, date_sent_to_procurement)
                            WHERE program_id = %s
                            """,
                            (rfp_shorthand, pm_name, agree_pm, dar_name,
                             sec_level, date_proc, program_id),
                        )

                # Link firm's categories to the program
                for cat_key in ["category_1", "category_2", "category_3", "category_4"]:
                    cat_name = cell_val(row, DATA_COL[cat_key])
                    if cat_name and rfp_number:
                        cat_id = get_or_create_category(cursor, cat_name)
                        cursor.execute(
                            "SELECT program_id FROM programs WHERE rfp_number = %s",
                            (rfp_number,),
                        )
                        prog_row = cursor.fetchone()
                        if cat_id and prog_row:
                            sort = int(cat_key[-1])  # 1, 2, 3, 4
                            cursor.execute(
                                """
                                INSERT IGNORE INTO program_categories
                                    (program_id, category_id, sort_order)
                                VALUES (%s, %s, %s)
                                """,
                                (prog_row[0], cat_id, sort),
                            )

            else:
                # Dry run — just count
                summary.inserted += 1

        except Exception as e:
            summary.errors.append(f"Row {row_idx} ({firm_name}): {e}")
            log.error(f"  ✗ Row {row_idx}: {e}")
            continue


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Import ECMS Excel data into MySQL")
    parser.add_argument("--file",    required=True, help="Path to the Excel file")
    parser.add_argument("--dry-run", action="store_true", help="Parse only, no DB writes")
    parser.add_argument("--verbose", action="store_true", help="Print every row processed")
    args = parser.parse_args()

    if not os.path.exists(args.file):
        log.error(f"File not found: {args.file}")
        sys.exit(1)

    mode = "DRY RUN" if args.dry_run else "LIVE"
    log.info(f"{'='*60}")
    log.info(f"  ECMS Excel Import — {mode}")
    log.info(f"  File : {args.file}")
    log.info(f"  Time : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log.info(f"{'='*60}")

    # Open workbook
    try:
        wb = openpyxl.load_workbook(args.file, data_only=True)
    except Exception as e:
        log.error(f"Could not open Excel file: {e}")
        sys.exit(1)

    # Detect sheets — flexible naming
    summary_sheet = None
    data_sheet     = None
    for name in wb.sheetnames:
        nl = name.lower()
        if "summary" in nl:
            summary_sheet = wb[name]
        elif "data" in nl:
            data_sheet = wb[name]

    if not summary_sheet:
        log.warning("Could not find a 'Summary' sheet — skipping Summary Level import")
    if not data_sheet:
        log.warning("Could not find a 'Data' sheet — skipping Data Level import")

    # Connect to DB
    conn = None
    cursor = None
    if not args.dry_run:
        try:
            conn = get_connection()
            cursor = conn.cursor()
            log.info("Connected to MySQL database")
        except Exception as e:
            log.error(f"Database connection failed: {e}")
            log.error("Check your .env file credentials")
            sys.exit(1)

    # --- Process Summary Level ---
    sum1 = RunSummary(sheet="Summary Level")
    if summary_sheet:
        import_summary_level(summary_sheet, cursor, sum1, args.dry_run, args.verbose)
        if not args.dry_run and conn:
            conn.commit()
    sum1.report()

    # --- Process Data Level ---
    sum2 = RunSummary(sheet="Data Level")
    if data_sheet:
        import_data_level(data_sheet, cursor, sum2, args.dry_run, args.verbose)
        if not args.dry_run and conn:
            conn.commit()
    sum2.report()

    # --- Final summary ---
    total_errors = len(sum1.errors) + len(sum2.errors)
    log.info("IMPORT COMPLETE")
    log.info(f"  Total rows read    : {sum1.rows_read + sum2.rows_read}")
    log.info(f"  Total inserted     : {sum1.inserted + sum2.inserted}")
    log.info(f"  Total updated      : {sum1.updated + sum2.updated}")
    log.info(f"  Total skipped      : {sum1.skipped + sum2.skipped}")
    log.info(f"  Total errors       : {total_errors}")
    log.info(f"  Log file           : {log_filename}")

    if cursor:
        cursor.close()
    if conn:
        conn.close()

    sys.exit(1 if total_errors > 0 else 0)


if __name__ == "__main__":
    main()